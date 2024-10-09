import os
import subprocess
import json
import csv
import yaml
from openpyxl import Workbook

def clone_repo(github_url, branch_name, target_folder):
    json_path = ''
    # Check if the target folder already exists
    if not os.path.exists(target_folder):
        os.makedirs(target_folder)
    else:
        print(f"Error: The folder '{target_folder}' already exists.")
    
    # Clone the specified branch into the target folder
    try:
        subprocess.run(
            ["git", "clone", "--branch", branch_name, github_url, target_folder],
            check=True
        )
        print(f"Successfully cloned '{branch_name}' branch into '{target_folder}'.")
    except subprocess.CalledProcessError as e:
        print(f"Error cloning the repository: {e}")


def fetch_json(target_folder):
    for root, folders, files in os.walk(target_folder):
        if "dev-values" in folders:  # Check if 'dev-values' is in the list of folders
            dev_values_path = os.path.join(root, "dev-values")  # Full path to 'dev-values'
            for filename in os.listdir(dev_values_path):  # List files in 'dev-values'
                if filename == "config-dev.json":  # Look for 'config-dev.json'
                    json_path = os.path.join(dev_values_path, filename)
                    break  # Exit after finding the file
        
        if json_path:  # If the JSON file is found, exit the loop
            break

    return json_path



def yaml_to_json(folder_path):
    json_data = {}

    # Convert YAML files to JSON structure
    for filename in os.listdir(folder_path):
        if filename.endswith('.yaml'):
            yaml_path = os.path.join(folder_path, filename)
            with open(yaml_path, 'r') as yaml_file:
                yaml_content = yaml.safe_load(yaml_file)
                json_data[filename[:-5]] = yaml_content  # Use filename without .yaml as the root key

    return json_data


def compare_json_files(old_data, new_data):
    changes = []

    # Check for changes in the JSON files
    for root in new_data.keys():
        if root not in old_data:
            changes.append((root, 'add', '', json.dumps(new_data[root], indent=4), 'Root object added'))
        else:
            compare(old_data[root], new_data[root], root, changes)

    for root in old_data.keys():
        if root not in new_data:
            changes.append((root, 'delete', '', '', 'Root object deleted'))

    return changes

def compare(old, new, root, changes, path=''):
    if isinstance(old, dict):
        for key in old.keys():
            new_key_path = f"{path}/{key}" if path else key
            if key in new:
                compare(old[key], new[key], root, changes, new_key_path)
            else:
                changes.append((root, 'delete', new_key_path, json.dumps(old[key], indent=4), 'Deleted'))

        for key in new.keys():
            new_key_path = f"{path}/{key}" if path else key
            if key not in old:
                changes.append((root, 'add', new_key_path, json.dumps(new[key], indent=4), 'Added'))

    elif isinstance(old, list):
        for item in old:
            if item not in new:
                changes.append((root, 'delete', path, json.dumps(item, indent=4), 'Deleted'))

    else:
        if old != new:
            changes.append((root, 'modify', path, json.dumps(new, indent=4), 'Modified'))

def write_changes_to_csv(changes, csv_file_path):
    with open(csv_file_path, 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        csv_writer.writerow(['Service name', 'Change Request', 'Key', 'Value', 'Comment'])

        for change in changes:
            service_name, change_type, key, value, comment = change
            csv_writer.writerow([service_name, change_type, key, value, comment])

def write_changes_to_excel(changes, excel_file_path, envs):
    # Create a new workbook and add sheets
    wb = Workbook()

    # Write the first sheet with full data
    first_sheet = wb.active
    first_sheet.title = envs[0]  # Name the first sheet

    # Write the headers
    first_sheet.append(['Service name', 'Change Request', 'Key', 'Value', 'Comment'])

    # Write the changes to the first sheet
    for change in changes:
        service_name, change_type, key, value, comment = change
        first_sheet.append([service_name, change_type, key, value, comment])

    # Write the remaining sheets (copy content but clear "Value" column)
    for env in envs[1:]:
        new_sheet = wb.create_sheet(title=env)
        new_sheet.append(['Service name', 'Change Request', 'Key', 'Value', 'Comment'])

        for row in first_sheet.iter_rows(min_row=2, values_only=True):
            # Copy all rows but clear the "Value" column (4th column)
            new_sheet.append([row[0], row[1], row[2], '', row[4]])

    # Save the workbook to a file
    wb.save(excel_file_path)


def get_input(prompt):
    attempts = 3
    for _ in range(attempts):
        value = input(prompt)
        if value:
            return value
        print("Input cannot be empty. Please try again.")
    print("Cannot proceed with execution as no input was entered.")
    exit(1)

def main():
    folder_path = get_input("Enter the path of the folder containing YAML files: ")
    csv_file_path = get_input("Enter the path to save the changes CSV file: ")
    excel_file_path = get_input("Enter the path to save the Excel file: ")
    github_url = get_input("Add the Repository URL: ")
    promote_branch_x-1 = get_input("Add the previous branch name : ") 
    promote_branch_x = get_input("Add the current branch name : ") 
    target_folder_x-1 = get_input("Add the folder where the x-1 repo has to be cloned : ")  
    target_folder_x = get_input("Add the folder where the x repo has to be cloned : ") 

    # Environment list (for this example, you can modify based on your use case)
    envs = ['dev', 'sit', 'uat']  # Example environments

    clone_repo(github_url, promote_branch_x-1, target_folder_x-1)

    previous_json_path = fetch_json(target_folder_x-1)
    print(previous_json_path)

    clone_repo(github_url, promote_branch_x, target_folder_x)

    new_json_path = fetch_json(target_folder_x)
    print(new_json_path)

    # Load previous JSON data
    try:
        with open(previous_json_path, 'r') as old_file:
            old_data = json.load(old_file)
    except (json.JSONDecodeError, FileNotFoundError):
        print("Previous JSON file is invalid or not found.")
        old_data = {}

    # Convert YAML to JSON
    json_data = yaml_to_json(folder_path)

    # Save the new JSON data
    with open(new_json_path, 'w') as new_json_file:
        json.dump(json_data, new_json_file, indent=4)



    # Compare JSON files
    changes = compare_json_files(old_data, json_data)

    # Write changes to CSV
    write_changes_to_csv(changes, csv_file_path)

    # Write changes to Excel with multiple sheets
    write_changes_to_excel(changes, excel_file_path, envs)

if __name__ == '__main__':
    main()
