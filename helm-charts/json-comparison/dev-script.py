import os
import json
import csv
import yaml
from openpyxl import Workbook

def yaml_to_json(folder_path):
    json_data = {}

    # Convert YAML files to JSON structure
    for filename in os.listdir(folder_path):
        if filename.endswith('.yaml'):
            yaml_path = os.path.join(folder_path, filename)
            with open(yaml_path, 'r') as yaml_file:
                yaml_content = yaml.safe_load(yaml_file)
                json_data[filename[:-5]] = yaml_content  # Use filename without .yaml as the root key

    return json_data


def compare_json_files(old_data, new_data):
    changes = []

    # Check for changes in the JSON files
    for root in new_data.keys():
        if root not in old_data:
            changes.append((root, 'add', '', json.dumps(new_data[root], indent=4), 'Root object added'))
        else:
            compare(old_data[root], new_data[root], root, changes)

    for root in old_data.keys():
        if root not in new_data:
            changes.append((root, 'delete', '', '', 'Root object deleted'))

    return changes

def compare(old, new, root, changes, path=''):
    if isinstance(old, dict):
        for key in old.keys():
            new_key_path = f"{path}/{key}" if path else key
            if key in new:
                compare(old[key], new[key], root, changes, new_key_path)
            else:
                changes.append((root, 'delete', new_key_path, json.dumps(old[key], indent=4), 'Deleted'))

        for key in new.keys():
            new_key_path = f"{path}/{key}" if path else key
            if key not in old:
                changes.append((root, 'add', new_key_path, json.dumps(new[key], indent=4), 'Added'))

    elif isinstance(old, list):
        # Check for modifications and deletions
        for i, old_item in enumerate(old):
            if i < len(new):
                if old_item != new[i]:
                    changes.append((root, 'modify', path, json.dumps(new[i], indent=4), 'Modified'))
            else:
                changes.append((root, 'delete', path, json.dumps(old_item, indent=4), 'Deleted'))

        # Check for additions
        for i in range(len(old), len(new)):
            changes.append((root, 'add', path, json.dumps(new[i], indent=4), 'Added'))

    else:
        if old != new:
            changes.append((root, 'modify', path, json.dumps(new, indent=4), 'Modified'))

def write_changes_to_csv(changes, csv_file_path):
    with open(csv_file_path, 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        csv_writer.writerow(['Service name', 'Change Request', 'Key', 'Value', 'Comment'])

        for change in changes:
            service_name, change_type, key, value, comment = change
            csv_writer.writerow([service_name, change_type, key, value, comment])

def write_changes_to_excel(changes, excel_file_path, envs):
    # Create a new workbook and add sheets
    wb = Workbook()

    # Write the first sheet with full data
    first_sheet = wb.active
    first_sheet.title = envs[0]  # Name the first sheet

    # Write the headers
    first_sheet.append(['Service name', 'Change Request', 'Key', 'Value', 'Comment'])

    # Write the changes to the first sheet
    for change in changes:
        service_name, change_type, key, value, comment = change
        first_sheet.append([service_name, change_type, key, value, comment])

    # Write the remaining sheets (copy content but clear "Value" column)
    for env in envs[1:]:
        new_sheet = wb.create_sheet(title=env)
        new_sheet.append(['Service name', 'Change Request', 'Key', 'Value', 'Comment'])

        for row in first_sheet.iter_rows(min_row=2, values_only=True):
            # Copy all rows but clear the "Value" column (4th column)
            new_sheet.append([row[0], row[1], row[2], '', row[4]])

    # Save the workbook to a file
    wb.save(excel_file_path)

def get_input(prompt):
    attempts = 3
    for _ in range(attempts):
        value = input(prompt)
        if value:
            return value
        print("Input cannot be empty. Please try again.")
    print("Cannot proceed with execution as no input was entered.")
    exit(1)

def main():
    folder_path = get_input("Enter the path of the folder containing YAML files: ")
    previous_json_path = get_input("Enter the path of the previous JSON file: ")
    new_json_path = get_input("Enter the path to save the new JSON file: ")
    csv_file_path = get_input("Enter the path to save the changes CSV file: ")
    excel_file_path = get_input("Enter the path to save the Excel file: ")

    # Environment list (for this example, you can modify based on your use case)
    envs = ['dev', 'sit', 'uat']  # Example environments

    # Convert YAML to JSON
    json_data = yaml_to_json(folder_path)

    # Save the new JSON data
    with open(new_json_path, 'w') as new_json_file:
        json.dump(json_data, new_json_file, indent=4)

    # Load previous JSON data
    try:
        with open(previous_json_path, 'r') as old_file:
            old_data = json.load(old_file)
    except (json.JSONDecodeError, FileNotFoundError):
        print("Previous JSON file is invalid or not found.")
        old_data = {}

    # Compare JSON files
    changes = compare_json_files(old_data, json_data)

    # Write changes to CSV
    write_changes_to_csv(changes, csv_file_path)

    # Write changes to Excel with multiple sheets
    write_changes_to_excel(changes, excel_file_path, envs)

if __name__ == '__main__':
    main()
