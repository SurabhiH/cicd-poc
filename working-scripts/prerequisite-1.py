import os
import git
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
 
def clone_repo(repo_path, clone_dir):
    if not os.path.exists(clone_dir):
        print(f"Cloning repository from {repo_path} into {clone_dir}...")
        git.Repo.clone_from(repo_path, clone_dir)
        print(f"Repository cloned successfully into {clone_dir}")
    else:
        print(f"Repository already exists at {clone_dir}, skipping cloning.")
 
# Function to read branches from the branch txt file
def read_branches(file_path):
    with open(file_path, 'r') as f:
        branches = [line.strip() for line in f.readlines()]
    return branches
 
# Function to get the list of files in the folder ending with 'values'
def get_files_in_values_folder(repo, branch):
    repo.git.checkout(branch)
    folder_path = ''
    for root, dirs, _ in os.walk(repo.working_dir):
        if root.endswith('values'):
            folder_path = root
            break
    if folder_path:
        return os.listdir(folder_path)
    return []
 
# Function to write data into Excel with unique values highlighting
def write_to_excel(repo_name, branch_files, output_file, wb=None):
    if wb is None:
        wb = openpyxl.Workbook()
 
    ws = wb.create_sheet(title=repo_name)
 
    # Write branch names to the headers
    for col_num, branch in enumerate(branch_files.keys(), start=1):
        ws.cell(row=1, column=col_num).value = branch
 
    # Write file names under each branch
    max_rows = max(len(files) for files in branch_files.values())
    for col_num, (branch, files) in enumerate(branch_files.items(), start=1):
        for row_num, file_name in enumerate(files, start=2):
            ws.cell(row=row_num, column=col_num).value = file_name
 
    # Collect unique values from the first column
    first_col_values = set()
    for row in range(2, max_rows + 2):  # Start from the second row
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is not None:
            first_col_values.add(cell_value)
 
    # Highlight unique values compared to the first column
    for col in range(2, len(branch_files) + 1):  # Start from the second column
        for row in range(2, max_rows + 2):  # Start from the second row
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None and cell_value not in first_col_values:
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
 
    # Save Excel file at the end
    wb.save(output_file)
    return wb
 
# New function to populate missing values from the first column in the rest of the columns
def populate_missing_values(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        # Collect values from the first column, excluding headers and empty cells
        first_col_values = [ws.cell(row=row, column=1).value for row in range(2, max_row + 1) if ws.cell(row=row, column=1).value is not None]
 
        # Check each value from the first column against the rest of the columns
        for value in first_col_values:
            for col in range(2, ws.max_column + 1):
                found = False
                for row in range(2, max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value == value:
                        found = True
                        break
                # If the value is not found, add it to the next available row without gaps
                if not found:
                    # Determine the next available row in the current column without leaving gaps
                    next_row = 2  # Start from the second row
                    while next_row <= max_row and ws.cell(row=next_row, column=col).value is not None:
                        next_row += 1
 
                    # If there's an empty cell, fill it; otherwise, append at the end
                    if next_row <= max_row:
                        ws.cell(row=next_row, column=col).value = value
                    else:
                        ws.cell(row=max_row + 1, column=col).value = value
                    ws.cell(next_row if next_row <= max_row else max_row + 1, column=col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
 
    wb.save(excel_file)
    print(f"Missing values populated in {excel_file}")


# Main function to process multiple repos
def main():
    repos_info = {
        'cs-helm-charats': r'https://github.hdfcbank.com/HDFCBANK/cs-helm-charats.git',
        'mb-helmcharts': r'https://github.hdfcbank.com/HDFCBANK/mb-helmcharts.git',
        'admin-helm-charts': r'https://github.hdfcbank.com/HDFCBANK/admin-helm-charts.git'
    }
    base_clone_dir = '/Users/mgxr2439/cicd-workspace/'
    output_file = 'pre-requisite.xlsx'
 
    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
 
    for repo_name, repo_url in repos_info.items():
        print(f"Processing repository: {repo_name}")
        branch_file = f'{repo_name}.txt'  # Assuming txt file is named after the repo
        clone_dir = os.path.join(base_clone_dir, repo_name)
 
        # Clone the repository
        clone_repo(repo_url, clone_dir)
        repo = git.Repo(clone_dir)
 
        # Read branch names
        branches = read_branches(branch_file)
 
        branch_files = {}
        for branch in branches:
            print(f'Processing branch: {branch} for repo {repo_name}')
            try:
                files = get_files_in_values_folder(repo, branch)
                branch_files[branch] = files
            except Exception as e:
                print(f'Error processing branch {branch} for repo {repo_name}: {e}')
                branch_files[branch] = []
 
        # Write the result to Excel
        wb = write_to_excel(repo_name, branch_files, output_file, wb=wb)
 
    print(f"Excel file saved: {output_file}")
 
    # Call the new function to populate missing values
    populate_missing_values(output_file)
 
if __name__ == "__main__":
    main()
