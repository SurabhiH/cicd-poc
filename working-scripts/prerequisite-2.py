import os
import json
import yaml
import subprocess
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define the repositories and their branch files
repositories = {
    'repo1': 'repo1.txt',
    'repo2': 'repo2.txt'
}

# Function to extract 'env' section from a YAML file
def extract_env_section(yaml_file):
    with open(yaml_file, 'r') as file:
        content = yaml.safe_load(file)
    
    env_section = {}
    
    if 'env' in content:
        env_section = content['env']
        
        # Filter to keep only the entries that have the 'name' key
        filtered_env = {
            key: [
                {'name': item['name']} for item in value if 'name' in item
            ]
            for key, value in env_section.items()
        }
        
        return filtered_env

    return env_section 

# Function to clone repos and process branches
def process_repositories():
    for repo_name, branches_file in repositories.items():
        # Clone the repository
        subprocess.run(['git', 'clone', f'https://github.com/SurabhiH/{repo_name}.git'], check=True)

        # Read the branches from the respective file
        with open(branches_file, 'r') as file:
            branches = file.read().splitlines()

        # Prepare output directory
        output_dir = os.path.join('output', repo_name)
        os.makedirs(output_dir, exist_ok=True)

        # Process each branch
        for branch in branches:
            # Switch to the branch
            subprocess.run(['git', 'checkout', branch], cwd=repo_name, check=True)

            helm_charts_dir = os.path.join(repo_name, 'helm-charts')
            branch_data = {}

            # Find all .yaml files in the helm-charts directory
            for root, _, files in os.walk(helm_charts_dir):
                for file in files:
                    if file.endswith('.yaml'):
                        yaml_file_path = os.path.join(root, file)
                        env_data = extract_env_section(yaml_file_path)
                        
                        # Use the file name (without extension) as the root object
                        branch_data[os.path.splitext(file)[0]] = env_data

            # Write the consolidated data to a single JSON file for the branch
            json_file_path = os.path.join(output_dir, f'{branch}.json')
            with open(json_file_path, 'w') as json_file:
                json.dump(branch_data, json_file, indent=4)

        # Return to the main branch (assumed to be 'main' or 'master')
        subprocess.run(['git', 'checkout', 'main'], cwd=repo_name, check=True)

def create_excel_report():
    output_base_dir = 'output'
    excel_path = os.path.join(output_base_dir, 'repositories_report.xlsx')
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')

    # Loop through each repository folder
    for repo in os.listdir(output_base_dir):
        repo_dir = os.path.join(output_base_dir, repo)
        
        if os.path.isdir(repo_dir):
            # Prepare data for the Excel sheet
            data = {}
            for json_file in os.listdir(repo_dir):
                if json_file.endswith('.json'):
                    json_path = os.path.join(repo_dir, json_file)
                    with open(json_path, 'r') as f:
                        branch_data = json.load(f)

                        # Format the data for each branch
                        for key, value in branch_data.items():
                            if value:  # Only add non-empty values
                                formatted_entry = f"{key}: {json.dumps(value)}"
                                if json_file not in data:
                                    data[json_file] = {}
                                data[json_file][key] = formatted_entry
            
            # Create a DataFrame for the current repository
            df = pd.DataFrame(data)

            # Write the DataFrame to a new sheet in the Excel file
            df.to_excel(writer, sheet_name=repo, index=False)

    # Save the Excel file
    writer.close()
    print(f"Excel report generated: {excel_path}")

def format_excel_report():
    excel_path = os.path.join('output', 'repositories_report.xlsx')
    workbook = load_workbook(excel_path)

    # Define fill colors
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        
        # Get the max row and column
        max_row = ws.max_row
        max_col = ws.max_column

        # Collect the root objects and their "name" keys from the first column
        first_col_data = {}
        for row in range(2, max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                try:
                    root_object, json_data = cell_value.split(": ", 1)
                    json_dict = json.loads(json_data)
                    # Collect all "name" values from all keys
                    first_col_data[root_object] = {item['name'] for key in json_dict for item in json_dict[key] if isinstance(item, dict) and 'name' in item}
                except Exception as e:
                    print(f"Error parsing JSON in row {row}: {e}")

        # Compare each column with the first column
        for col in range(2, max_col + 1):
            column_data = {}
            for row in range(2, max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    try:
                        root_object, json_data = cell_value.split(": ", 1)
                        json_dict = json.loads(json_data)
                        # Collect all "name" values from all keys
                        column_data[root_object] = {item['name'] for key in json_dict for item in json_dict[key] if isinstance(item, dict) and 'name' in item}
                    except Exception as e:
                        print(f"Error parsing JSON in column {col}, row {row}: {e}")

            next_empty_row = max_row + 1

            # Check for mismatches and missing "name" values
            for root_object, first_names in first_col_data.items():
                if root_object in column_data:
                    column_names = column_data[root_object]
                    # Check for names that are in the first column but not in the current column
                    for name in first_names:
                        if name not in column_names:
                            # Append the missing name in the next available row in the current column
                            ws.cell(row=next_empty_row, column=col).value = f"{root_object}: {{'name': '{name}'}}"
                            ws.cell(row=next_empty_row, column=col).fill = red_fill
                            next_empty_row += 1  # Move to the next row for the next missing value
                else:
                    # If the root object does not exist in the column, append all names
                    for name in first_names:
                        ws.cell(row=next_empty_row, column=col).value = f"{root_object}: {{'name': '{name}'}}"
                        ws.cell(row=next_empty_row, column=col).fill = red_fill
                        next_empty_row += 1

            # Check for values in the current column that are not in the first column (yellow highlight)
            for col in range(2, max_col + 1):
                for row in range(2, max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        try:
                            root_object, json_data = cell_value.split(": ", 1)
                            json_dict = json.loads(json_data)
                            column_names = {item['name'] for key in json_dict for item in json_dict[key] if isinstance(item, dict) and 'name' in item}
                            
                            # Check if this name is in the first column's names
                            if root_object in first_col_data:
                                first_names = first_col_data[root_object]
                                for name in column_names:
                                    if name not in first_names:
                                        ws.cell(row=row, column=col).fill = yellow_fill
                                        break  # Highlight only once for each cell
                        except Exception as e:
                            print(f"Error parsing JSON in column {col}, row {row}: {e}")


    # Save the modified Excel file
    workbook.save(excel_path)
    print(f"Excel report formatted: {excel_path}")

if __name__ == '__main__':
    process_repositories()
    create_excel_report()
    format_excel_report()
