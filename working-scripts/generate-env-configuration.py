import os
import re
import yaml
import json
import pandas as pd
import openpyxl
 
 
def read_yaml_files_to_json(folder_path):
    json_data = {}
 
    for filename in os.listdir(folder_path):
        if filename.endswith('.yaml') or filename.endswith('.yml'):
            yaml_file_path = os.path.join(folder_path, filename)
            with open(yaml_file_path, 'r') as f:
                yaml_content = yaml.safe_load(f)
                root_object = os.path.splitext(filename)[0]
                json_data[root_object] = yaml_content
    return json_data
 
def try_parse_json(value):
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return value
 
def handle_data_env(json_data, service_name, change_request, parsed_value):
    """Handles modifications for 'data' and 'env' services."""
    obj = json_data.setdefault(service_name, [])
    
    if change_request == 'add':
        if not any(entry['name'] == parsed_value['name'] for entry in obj):
            obj.append(parsed_value)
            print(f"Added to '{service_name}': {parsed_value}")
        else:
            print(f"Warning: Entry with name '{parsed_value['name']}' already exists in '{service_name}'.")
 
    elif change_request == 'modify':
        for entry in obj:
            if entry['name'] == parsed_value['name']:
                entry.update(parsed_value)
                print(f"Modified entry in '{service_name}': {entry}")
                break
        else:
            print(f"Warning: Attempted to modify non-existent entry '{parsed_value['name']}' in '{service_name}'.")
 
    elif change_request == 'delete':
        json_data[service_name] = [entry for entry in obj if entry['name'] != parsed_value['name']]
        print(f"Deleted entry from '{service_name}' with name '{parsed_value['name']}'.")
 
def insert_hardcoded_value(text, service_file):
    """
    Inserts a hardcoded multi-line value in the 'env' section at the last-but-one position
    if multiple values exist, or at the end if only one value is present.
    """
 
    # Regex to locate the 'env:' section and capture its contents
    env_pattern = r"(env:\s*\n(?:\s*\{\{.*\}\}\n?)+)"
    # Find the env block
    match = re.search(env_pattern, text)
    if not match:
        print("No 'env' section found.")
        return text
 
    # Extract the env block and split it into lines
    env_block = match.group(1)
    env_lines = env_block.strip().splitlines()
 
    if env_lines[-1]:
        indent = len(env_lines[-1]) - len((env_lines[-1]).strip())
        indentation = " " * indent
        print(indentation + ":indent")
 
    # Define the hardcoded block using the `service_file` variable
    hardcoded_block = f"""{{{{- with .Values.env.{service_file} }}}}
              {{{{- toYaml . | nindent 12 }}}}
            {{{{- end}}}}"""
 
 
    # Insert the hardcoded block at the last-but-one position, if possible
    env_lines.append(indentation + hardcoded_block.strip())
    env_lines.append('\n')
 
    # Reassemble the modified env block and replace it in the original text
    modified_env_block = "\n".join(env_lines)
    modified_text = text.replace(env_block, modified_env_block)
 
    return modified_text
 
def update_template(service_file):
    temp_path = rf'/Users/mgxr3734/Desktop/generate-config/promotion-x/cs-helm-charats/helm-charts/templates/deployment.yaml'
   
    # Read the YAML file as text
    with open(temp_path, 'r') as file:
        text = file.read()
 
    # Insert the hardcoded value
    modified_text = insert_hardcoded_value(text, service_file)
 
    # Write the modified content back to the YAML file
    with open(temp_path, 'w') as file:
        file.write(modified_text)
 
 
def apply_changes_to_json(json_data, excel_file_path, sheet_name):
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
    print(excel_file_path)
    for index, row in df.iterrows():
        service_name = row['Service name']
        change_request = row['Change Request']
        key = row['Key']
        comments = row['Comment']
        value = row['Value']
 
        if pd.isna(value) or value == "":
            if change_request == "delete":
                continue
            else:
                raise ValueError(f"Missing or empty value encountered in row {index + 1}.")
 
        # Handle addition of a root object
        if pd.isna(key) or key == "":
            if comments == "root object added":
                new_root_object = service_name
                if new_root_object not in json_data:
                    json_data[new_root_object] = {}
                    print(f"Added new root object: {new_root_object}")
                if not pd.isna(value) and value != "":
                    json_data[new_root_object] = try_parse_json(value)
                    print(f"Stored value in '{new_root_object}': {value}")
                update_template(service_name)
 
                continue
            elif comments == "root object deleted":
                if service_name in json_data:
                    del json_data[service_name]
                    print(f"Deleted root object: {service_name}")
                continue
            elif service_name in ['data', 'env']:
                pass # Skip the key checks for 'data' and 'env'
            else:
                print(f"Error: Key is missing or empty in row {index + 1}.")
                raise ValueError(f"Missing or empty key encountered in row {index + 1}.")
 
        parsed_value = try_parse_json(value)
        if service_name in ['data', 'env']:
            handle_data_env(json_data, service_name, change_request, parsed_value)
            continue
 
        # General handling for other services with keys
        key_path = key.split('//')
        obj = json_data.setdefault(service_name, {})
 
        for k in key_path[:-1]:
            obj = obj.setdefault(k, {})
 
        final_key = key_path[-1]
        
        if key_path[0] == "env":
                if change_request == 'add':
                    # opts = obj[final_key].pop(-1)
                    if final_key not in obj:
                        obj[final_key] = []
                    obj[final_key].insert(-1,parsed_value)
                    print(f"Added to '{final_key}' in '{service_name}': {parsed_value}")
                    # obj[final_key].append(opts)
                elif change_request == 'modify':
                    for entry in obj[final_key]:
                        if entry['name'] == parsed_value['name']:
                            entry['value'] = parsed_value['value']
                            print(f"Modified '{final_key}' entry in '{service_name}': {parsed_value}")
                            break
                elif change_request == 'delete':
                    if final_key in obj:
                        obj[final_key] = [entry for entry in obj[final_key] if entry['name'] != parsed_value['name']]
                        print(f"Deleted entry from '{final_key}' in '{service_name}'.")
 
        else:
            if change_request == 'modify':
                obj[final_key] = parsed_value
                print(f"Modified '{final_key}' in '{service_name}': {parsed_value}")
            elif change_request == 'add':
                obj[final_key] = parsed_value
                print(f"Added '{final_key}' in '{service_name}': {parsed_value}")
            elif change_request == 'delete' and final_key in obj:
                del obj[final_key]
                print(f"Deleted '{final_key}' from '{service_name}'.")
 
    return json_data
 
def save_json_to_file(json_data, output_file):
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    with open(output_file, 'w') as f:
        json.dump(json_data, f, indent=4)
 
def process_json_data(data):
    if isinstance(data, list):
        return [process_json_data(item) for item in data]
    elif isinstance(data, dict):
        return {key: process_json_data(value) for key, value in data.items()}
    return data
 
def create_yaml_files_from_json(updated_output_file, output_folder):
    with open(updated_output_file, 'r') as json_file:
        json_data = json.load(json_file)
 
    os.makedirs(output_folder, exist_ok=True)
 
    for root_object, data in json_data.items():
        yaml_file_path = os.path.join(output_folder, f"{root_object}.yaml")
        processed_data = process_json_data(data)
        
        with open(yaml_file_path, 'w') as yaml_file:
            yaml.dump(processed_data, yaml_file, default_flow_style=False, sort_keys=False)
 
def check_updated_env(excel_file):
    wb = pd.ExcelFile(excel_file)
    sheets = []
    for sheet_name in wb.sheet_names[1:]:
        if sheet_name != "dev":
            df= pd.read_excel(wb, sheet_name=sheet_name)
            if "Value" in df.columns and df["Value"].notna().any():
                sheets.append(sheet_name)
 
    if sheets == []:
        print("No values updated in the release to promote the env")
    return sheets
 
def create_txt_file(excel_path,env,txt_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[env]  # Use sheet argument correctly
    max_row = ws.max_row
 
    # Collect unique values from the first column
    first_col_values = set()
    for row in range(2, max_row + 1):  # Start from the second row
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is not None:
            first_col_values.add(cell_value)
 
    with open(txt_path, 'w') as file:
        for name in first_col_values:
            if name =='data' or name =='env':
                pass
            else:
                file.write(f"{name}\n")    
        
def main():
    release_note_path = rf'/Users/mgxr3734/Desktop/generate-config/promotion-x/cs-helm-charats/release_note'
    repo_x = rf'/Users/mgxr3734/Desktop/generate-config/promotion-x/cs-helm-charats'
    for foldername in os.listdir(repo_x):
        if foldername.endswith('helm-charts') and os.path.exists(f"{release_note_path}/release-note.xlsx"):
            excel_file_path =  os.path.join(release_note_path,f"release-note.xlsx")
            sheet_name = check_updated_env(excel_file_path)
            for sheet in sheet_name:
                print(f"Promoting the values in env: {sheet} of {foldername}")
                folder_path = rf'/Users/mgxr3734/Desktop/generate-config/promotion-x-1/cs-helm-charats/helm-charts/{sheet}-values'
                initial_output_file = rf'/Users/mgxr3734/Desktop/generate-config/promotion-x-1/cs-helm-charats/helm-charts/{sheet}-values/config-{sheet}.json'
                updated_output_file = rf'/Users/mgxr3734/Desktop/generate-config/promotion-x/cs-helm-charats/helm-charts/{sheet}-values/config-{sheet}.json'
                output_folder = rf'/Users/mgxr3734/Desktop/generate-config/promotion-x/cs-helm-charats/helm-charts/{sheet}-values'
                txt_file_path = rf'/Users/mgxr3734/Desktop/generate-config/promotion-x/cs-helm-charats/helm-charts/{sheet}-values/{sheet}.txt'
                json_data = read_yaml_files_to_json(folder_path)
                save_json_to_file(json_data, initial_output_file)
                updated_json = apply_changes_to_json(json_data, excel_file_path, sheet)
                save_json_to_file(updated_json, updated_output_file)
                create_yaml_files_from_json(updated_output_file, output_folder)
                create_txt_file(excel_file_path,sheet,txt_file_path)
 
        elif foldername.endswith('helm-charats') and os.path.exists(f"release-note.xlsx") == False:
            print(f"]release-note.xlsx not present.")
            
if __name__ == "__main__":
    main()
