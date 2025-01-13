import os
import subprocess
import json
import yaml
from openpyxl import Workbook, load_workbook
from filecmp import cmp
 
def clone_repo(repo_url, branch_name, target_folder):
    try:
        os.makedirs(target_folder)  # Create the target folder
    except Exception as e:
        print(f"Error creating folder '{target_folder}': {e}")
        return
   
    # Clone the specified branch into the target folder
    try:
        subprocess.run(
            ["git", "clone", "--branch", branch_name, repo_url, target_folder],
            check=True
        )
        print(f"Successfully cloned '{branch_name}' branch into '{target_folder}'.")
    except subprocess.CalledProcessError as e:
        print(f"Error cloning the repository: {e}")
 
def fetch_json(target_folder, env):
    json_path = ''
    foldername = f"{env}-values"
    for root, folders, files in os.walk(target_folder):
        if foldername in folders:
            lower_env_path = os.path.join(root, foldername)
            for filename in os.listdir(lower_env_path):
                if filename == f"config-{env}.json":
                    json_path = os.path.join(lower_env_path, filename)
                    break
            if not json_path:
                # Create an empty config-dev.json if it doesn't exist
                json_path = os.path.join(lower_env_path, f"config-{env}.json")
                with open(json_path, 'w') as f:
                    json.dump({}, f)  # Create an empty JSON object
            break  # Exit after processing the 'dev-values' folder
 
    return json_path
 
 
def yaml_to_json(folder_path):
    json_data = {}
    # Convert YAML files to JSON structure
    for filename in os.listdir(folder_path):
        if filename.endswith('.yaml'):
            yaml_path = os.path.join(folder_path, filename)
            with open(yaml_path, 'r') as yaml_file:
                yaml_content = yaml.safe_load(yaml_file)
                json_data[filename[:-5]] = yaml_content  # Use filename without .yaml as the root key
 
    return json_data
 
def compare_json_files(old_data, new_data):
    changes = []
 
    # Check for changes in the JSON files
    for root in new_data.keys():
        if root not in old_data:
            changes.append((root, 'add', '', json.dumps(new_data[root], indent=4), '','root object added'))
        else:
            compare(old_data[root], new_data[root], root, changes)
 
    for root in old_data.keys():
        if root not in new_data:
            changes.append((root, 'delete', '', json.dumps(old_data[root], indent=4), '', 'root object deleted'))
 
    return changes
 
def compare(old, new, root, changes, path=''):
    # Handle dictionary structures
    if isinstance(old, dict) and isinstance(new, dict):
        for k in old.keys():
            new_key_path = f"{path}//{k}" if path else k
            if k in new:
                compare(old[k], new[k], root, changes, new_key_path)
            else:
                changes.append((root, 'delete', new_key_path, json.dumps(old[k], indent=4), '', 'Deleted'))
 
        for k in new.keys():
            new_key_path = f"{path}//{k}" if path else k
            if k not in old:
                changes.append((root, 'add', new_key_path, json.dumps(new[k], indent=4),'', 'Added'))
 
    # Handle lists
    elif isinstance(old, list) and isinstance(new, list):
        if all(isinstance(i, dict) for i in old) and all(isinstance(i, dict) for i in new):
            # Check if both lists are not empty
            if old and new and "name" in old[0] and "name" in new[0]:
                compare_list_of_dicts(old, new, root, changes, path)
            else:
                # Handle lists without "name" key or if lists are empty
                for i, old_item in enumerate(old):
                    if i < len(new):
                        if old_item != new[i]:
                            changes.append((root, 'modify', f"{path}", "["+json.dumps(new[i], indent=4)+"]", "["+json.dumps(old_item, indent=4)+"]",'Modified'))
                    else:
                        changes.append((root, 'delete', f"{path}", json.dumps(old_item, indent=4), '','Deleted'))
 
                # Add any new elements from the new list
                for i in range(len(old), len(new)):
                    changes.append((root, 'add', f"{path}", json.dumps(new[i], indent=4), '', 'Added'))
 
    # Compare scalar values
    else:
        if old != new:
            changes.append((root, 'modify', path, json.dumps(new, indent=4), json.dumps(old, indent=4) ,'Modified'))
 
def compare_list_of_dicts(old_list, new_list, root, changes, path=''):
    # Compare lists of dictionaries based on the "name" key
    old_dict = {item["name"]: item for item in old_list if "name" in item}
    new_dict = {item["name"]: item for item in new_list if "name" in item}
 
    # Compare old dictionaries with new
    for key, old_item in old_dict.items():
        if key in new_dict:
            new_item = new_dict[key]
            # If there are changes in the item
            if old_item != new_item:
                changes.append((root, 'modify', path, json.dumps(new_item, indent=4), json.dumps(old_item, indent=4),'Modified'))
        else:
            changes.append((root, 'delete', path, json.dumps(old_item, indent=4), '','Deleted'))
 
    # Check for additions
    for key, new_item in new_dict.items():
        if key not in old_dict:
            changes.append((root, 'add', path, json.dumps(new_item, indent=4),'', 'Added'))
 
 
def write_changes_to_excel(changes, release_note_path, envs):
    if not changes:
        print("No differences found, skipping the creation of release note.")
        return
    excel_file = "release-note.xlsx"
    excel_file_path = os.path.join(release_note_path, excel_file)
 
    # Check if the file exists
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
    else:
        wb = Workbook()
 
    first_sheet = wb.active
    first_sheet.title = envs[0]  # Name the first sheet
 
    first_sheet.append(['Service name', 'Change Request', 'Key', 'Value', 'Value before modification', 'Comment'])
 
        # Write the changes for the environment
    for change in changes:
        service_name, change_type, key, value, prev_value, comment = change
        # Append new changes to the relevant sheet
        first_sheet.append([service_name, change_type, key, value, prev_value, comment])
 
    for env in envs[1:]:
        new_sheet = wb.create_sheet(title=env)
        new_sheet.append(['Service name', 'Change Request', 'Key', 'Value', 'Value before modification','Comment'])
 
        for row in first_sheet.iter_rows(min_row=2, values_only=True):
            # Copy all rows but clear the "Value" column (4th column)
            new_sheet.append([row[0], row[1], row[2], '', row[4], row[5]])
 
    # Save the updated workbook
    wb.save(excel_file_path)
 
def parse_service_tags(file_path):
    """
    Parses the service tags from the 'update_image_tags.txt' file and returns a dictionary
    with service names as keys and build tags as values.
   
    Args:
        file_path (str): Path to the 'update_image_tags.txt' file.
   
    Returns:
        dict: Dictionary with service names as keys and build tags as values.
    """
    service_tags = {}
 
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()
 
            for line in lines:
                # Remove leading/trailing whitespaces
                line = line.strip()
                print(line)
               
                if not ':' in line:
                    continue
                service_name, build_tag = line.split(":",1)
                service_name = service_name.strip()
                build_tag = build_tag.strip()
                service_tags[service_name] = build_tag
 
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
    except Exception as e:
        print(f"Error while parsing the file '{file_path}': {e}")
    print(f"service_tags: {service_tags}")
    return service_tags
 
def update_image_tags_in_release_note(service_tags, release_note_path, envs):
    # Prepare changes for image tags
    image_changes = []
    for service_name, build_tag in service_tags.items():
        image_tag = str(build_tag).split(":")[-1]
        # Add the image tag modification to the changes list
        image_changes.append((
            service_name,
            'modify',
            'image//image_name',  # Assuming 'image_name' is the correct key
            build_tag,
            '',
            'Modified'
        ))
        image_changes.append((
            service_name,
            'modify',
            'image//tag',  # Assuming 'image_name' is the correct key
            image_tag,
            '',
            'Modified'
        ))
        print(image_tag)
 
    # Define the Excel file path
    excel_file = "release-note.xlsx"
    excel_file_path = os.path.join(release_note_path, excel_file)
 
    # Check if the file exists
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
    else:
        wb = Workbook()
 
    
    if envs[0] in wb.sheetnames:
        sheet = wb[envs[0]]
    else:
        ws = wb.create_sheet(title=env)
        ws.append(['Service name', 'Change Request', 'Key', 'Value', 'Value before modification', 'Comment'])
        sheet = ws
 
    # Append image changes
    for change in image_changes:
        service_name, change_type, key, value, prev_value, comment = change
        sheet.append([service_name, change_type, key, value, prev_value, comment])
 
    for env in envs[1:]:
        if env in wb.sheetnames:
            new_sheet = wb[env]
            for change in image_changes:
                service_name, change_type, key, value, prev_value, comment = change
                new_sheet.append([service_name, change_type, key, value, prev_value, comment])
        else:
            new_sheet = wb.create_sheet(title=env)
            new_sheet.append(['Service name', 'Change Request', 'Key', 'Value', 'Value before modification','Comment'])
            for change in image_changes:
                service_name, change_type, key, value, prev_value, comment = change
                new_sheet.append([service_name, change_type, key, value, prev_value, comment])
 
 
 
 
 
    # Save the updated workbook
    wb.save(excel_file_path)
 
    return image_changes
 
def get_input(prompt):
    attempts = 3
    for _ in range(attempts):
        value = input(prompt)
        if value:
            return value
        print("Input cannot be empty. Please try again.")
    print("Cannot proceed with execution as no input was entered.")
    exit(1)
 
def compare_shell_scripts(folder_x_1, folder_x, release_note_path, env):
    # Define paths to the scripts folders in each branch
    scripts_folder_x_1 = os.path.join(folder_x_1, f'helm-charts/{env}-values/db-scripts')
    scripts_folder_x = os.path.join(folder_x, f'helm-charts/{env}-values/db-scripts')
    script_txt_path = os.path.join(folder_x, f'helm-charts/scripts.txt')
    # Check if the scripts folders exist
    if not os.path.isdir(scripts_folder_x_1) or not os.path.isdir(scripts_folder_x):
        print("Scripts folder not found in one or both branches.")
        return
    # List of modified shell scripts
    modified_scripts = []
 
    # Compare files in both folders
    for filename in os.listdir(scripts_folder_x):
        if filename.endswith('.sh') and filename in os.listdir(scripts_folder_x_1):
            # Compare files with the same name in both folders
            file_path_x_1 = os.path.join(scripts_folder_x_1, filename)
            file_path_x = os.path.join(scripts_folder_x, filename)
            if not cmp(file_path_x_1, file_path_x, shallow=False):
                modified_scripts.append(filename)
 
    # If there are modified scripts, add a new sheet in the workbook
    if modified_scripts:
        excel_file = f"release-note.xlsx"
        excel_file_path = os.path.join(release_note_path, excel_file)
        if not os.path.exists(excel_file_path):
            wb = Workbook()
        else:
            wb = load_workbook(excel_file_path)
   
        ws = wb.create_sheet(title="db-scripts")
        # Add a new sheet for scripts if modified scripts are found
        ws.append(['Modified Shell Scripts'])
        # Write the list of modified scripts
        for script_name in modified_scripts:
            ws.append([script_name])
            with open(script_txt_path, 'a') as file:
                file.write(f"{script_name}\n")
            print(script_name)
        # Save the updated workbook
        wb.save(excel_file_path)
           
 
def main():
    repos_info = {
        'cs-helm-charats': r'https://github.hdfcbank.com/HDFCBANK/cs-helm-charats.git'
    }
 
    for repo_name, repo_url in repos_info.items():
        promote_branch_x_1 = input("Enter the branch containing the stable release: ")
        promote_branch_x = input("Enter the branch containing the updated files for release: ")
        target_folder_x_1 = rf'/Users/mgxr3734/Desktop/generate-config/promotion-x-1/{repo_name}'
        target_folder_x = rf'/Users/mgxr3734/Desktop/generate-config/promotion-x/{repo_name}'
 
 
    envs = input("Enter the envs to be promoted separated by a space: ").split()
    print(envs)
    
    # Environment list (you can modify based on your use case)
    # envs = ['dev', 'sit', 'uat', 'preprod', 'perf', 'mig/dm', 'sec', 'prod']  
 
    # To Clone the repo from promotion-x-1 branch
    clone_repo(repo_url, promote_branch_x_1, target_folder_x_1)
 
    # To Clone the repo from promotion-x branch
   
    clone_repo(repo_url, promote_branch_x, target_folder_x)
 
    #to create the release note folder
    release_note_path = os.path.join(target_folder_x, "release_note")
    if not os.path.exists(release_note_path):
        os.makedirs(release_note_path)
 
 
    #To fetch the path of config-dev.json from promotion-x-1 branch
    previous_json_path = fetch_json(target_folder_x_1, envs[0])
    print(previous_json_path)
 
    #To fetch the path of config-dev.json from promotion-x branch
    new_json_path = fetch_json(target_folder_x, envs[0])
    print(new_json_path)
 
    # Load previous JSON data
    try:
        with open(previous_json_path, 'r') as old_file:
            old_data = json.load(old_file)
    except (json.JSONDecodeError, FileNotFoundError):
        print("Previous JSON file is invalid or not found.")
        old_data = {}
 
    # Convert YAML to JSONhelm-charts/dev-values/upgrade-services.txt
    lower_env_path = os.path.dirname(new_json_path)
    json_data = yaml_to_json(lower_env_path)
 
    # Save the new JSON data
    with open(new_json_path, 'w') as new_json_file:
        json.dump(json_data, new_json_file, indent=4)
 
    # Compare JSON files
    changes = compare_json_files(old_data, json_data)
        # Write changes to Excel with multiple sheets
    write_changes_to_excel(changes, release_note_path, envs)
 
    txt_file_path = os.path.join(target_folder_x, f"helm-charts/{envs[0]}-values/upgrade-services.txt")
    print(f"image-promotion.txt file path: {txt_file_path}")
    service_tags = parse_service_tags(txt_file_path)
    if not service_tags:
        print("No service tags found or file was not read correctly.")
        return
   
    image_changes = update_image_tags_in_release_note(service_tags, release_note_path, envs)
    if image_changes:
        print(f"modified image names updated")
 
    # Compare shell scripts and add to release note if modified
    compare_shell_scripts(target_folder_x_1, target_folder_x, release_note_path, envs[0])
    target_folder_x = os.path.dirname(target_folder_x)
    target_folder_x_1 = os.path.dirname(target_folder_x_1)
 
if __name__ == '__main__':
    main()
