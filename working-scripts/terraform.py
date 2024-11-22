import openpyxl
import os
import datetime

# Helper function to read Excel sheet
def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = {}
    for col in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        column_name = col[0].value
        data[column_name] = [cell.value for cell in col[1:] if cell.value is not None]
    return data

# Helper function to load and replace placeholders in template files
def process_template(template_path, replacements):
    if not os.path.exists(template_path):
        return ""
    
    with open(template_path, 'r') as file:
        content = file.read()
    
    # Replace placeholders
    for placeholder, value in replacements.items():
        content = content.replace(placeholder, value)
    
    return content

# Function to generate the auto.tfvars file content
def generate_tfvars(data, templates_folder, global_file, output_file):
    current_date = datetime.datetime.now().strftime("%d-%b-%Y")
    
    # Read the global configurations first
    with open(global_file, 'r') as file:
        global_content = file.read()
    
    # Initialize the content as a dictionary
    tfvars_dict = {}
    
    # Loop through each resource (excluding 'ID' column)
    for column, values in data.items():
        if column.lower() == "id":  # Skip the 'ID' column
            continue
        
        # Create a dictionary for each column
        column_dict = {}
        
        # Process each value for the column
        for idx, value in enumerate(values):
            template_file = os.path.join(templates_folder, f"{column.lower()}.txt")
            if os.path.exists(template_file):
                replacements = {
                    "<<id>>": str(idx + 1),  # Index + 1 for ID
                    "<<name>>": value,
                    "<<current_date>>": current_date
                }
                section_content = process_template(template_file, replacements)
                column_dict[f"{column.lower()}_{idx + 1}"] = section_content
            else:
                print(f"Warning: Template for '{column}' not found.")
        
        # Store the column dictionary in the main tfvars dictionary
        tfvars_dict[column] = column_dict

    # Initialize the tfvars content with the global settings
    tfvars_content = global_content + "\n"
    
    # Write the data from the tfvars dictionary to the auto.tfvars content
    for column, column_dict in tfvars_dict.items():
        # Section header for the resource column
        section_begins = f"###################################### {column.upper()} ###########################################\n"
        section_begins += f"{column.lower()} = {{\n\n"  # Open the dictionary block for each section
        tfvars_content += section_begins
        
        # Add each value and its content to the tfvars_content
        for key, content in column_dict.items():
            tfvars_content += f"    {key} = {{\n{content}\n    }},\n\n"
        
        section_ends = "}\n\n"  # Close the dictionary block
        tfvars_content += section_ends
        tfvars_content+= f"###################################### {column.upper()} ENDS ###########################################\n"

    # Write the generated content to auto.tfvars
    with open(output_file, 'w') as file:
        file.write(tfvars_content)
    print(f"auto.tfvars file generated: {output_file}")

# Main function to run the script
def main():
    # Paths to the required files
    excel_file = "data.xlsx"  # Excel file with resource data
    templates_folder = "templates"  # Folder where template files are stored
    global_file = "templates/global.txt"  # File with global settings
    output_file = "auto.tfvars"  # Output auto.tfvars file

    # Step 1: Read data from the Excel sheet
    data = read_excel(excel_file)

    # Step 2: Generate the auto.tfvars file
    generate_tfvars(data, templates_folder, global_file, output_file)

# Run the script
if __name__ == "__main__":
    main()
